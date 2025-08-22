import xml.etree.ElementTree as ET
import os
import pandas as pd
from openpyxl import load_workbook
from collections import defaultdict
import shutil
import math
import re
import datetime

# Конфигурационные константы
CONFIG = {
    'MAX_ROWS_PER_SHEET': 52,
    'DATA_START_ROW': 4,  # строка начала данных (после заголовков)
    'TEMPLATE_FILE': 'A3.xlsx',
    'DEFAULT_SHEET_NAME': 'Л1',
    'COLUMN_MAPPING': {
        'InterfaceName': 'E',  # Наименование сигнала
        'Comment': 'AD',  # Смысловое значение сигнала
        'IoType': 'Q',  # Тип В/В
        'LogicalAddress': 'S'  # № канала
    },
    'COLUMN_DISPLAY_NAMES': {
        'InterfaceName': 'Наименование сигнала',
        'Comment': 'Смысловое значение сигнала',
        'IoType': 'Тип В/В',
        'LogicalAddress': '№ канала'
    }
}


def _validate_file_path(file_path, file_type="файл", check_exists=True):
    """Проверка валидности пути к файлу"""
    if not file_path or not file_path.strip():
        raise ValueError(f"Путь к {file_type} не может быть пустым")

    file_path = file_path.strip()

    # Проверка на запрещенные символы в имени файла
    forbidden_chars = ['/', '\\', ':', '*', '?', '"', '<', '>', '|']
    filename = os.path.basename(file_path)
    if any(char in filename for char in forbidden_chars):
        raise ValueError(f"Имя {file_type} содержит запрещенные символы")

    if check_exists and not os.path.exists(file_path):
        raise FileNotFoundError(f"{file_type.capitalize()} не найден: {file_path}")

    return os.path.abspath(file_path)


def _get_safe_input(prompt, validation_func=None, default=None):
    """Безопасный ввод с валидацией"""
    while True:
        try:
            value = input(prompt).strip()
            if not value and default is not None:
                return default
            if validation_func:
                return validation_func(value)
            return value
        except (KeyboardInterrupt, EOFError):
            print("\nПрервано пользователем")
            exit(1)
        except Exception as e:
            print(f"Ошибка: {e}. Попробуйте снова.")


def parse_aml_file(file_path):
    """Парсинг AML-файла с созданием отдельной строки для каждого ExternalInterface"""
    try:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Файл не найден: {file_path}")

        tree = ET.parse(file_path)
        root = tree.getroot()

        # Определяем namespace динамически
        namespace = {'caex': root.tag.split('}')[0][1:]} if '}' in root.tag else {'caex': ''}

        # Список для хранения всех строк данных
        rows = []

        # Проходим по всех устройствам
        for device_elem in root.findall('.//caex:InternalElement', namespace):
            device_name = device_elem.get('Name', 'N/A')

            # Пропускаем устройства без ключевых слов
            if not any(kw in device_name for kw in ["Rack", "Rail"]):
                continue

            device_type = device_elem.get('RefBaseClassPath', 'N/A').split('/')[-1] if device_elem.get(
                'RefBaseClassPath') else 'N/A'

            # Обрабатываем все интерфейсы устройства
            for interface in device_elem.findall('.//caex:ExternalInterface', namespace):
                row_data = defaultdict(lambda: "")
                row_data['DeviceName'] = device_name
                row_data['DeviceType'] = device_type
                row_data['InterfaceName'] = interface.get('Name', 'N/A')

                # Обрабатываем атрибуты интерфейса
                for attr in interface.findall('.//caex:Attribute', namespace):
                    attr_name = attr.get('Name', 'UnknownAttr')

                    # Получаем значение атрибута
                    value_elem = attr.find('caex:Value', namespace)
                    if value_elem is not None and value_elem.text is not None:
                        row_data[attr_name] = value_elem.text.strip()

                    # Обрабатываем вложенные атрибуты
                    for nested_attr in attr.findall('.//caex:Attribute', namespace):
                        nested_name = nested_attr.get('Name', 'NestedAttr')
                        full_nested_name = f"{attr_name}.{nested_name}"

                        nested_value = nested_attr.find('caex:Value', namespace)
                        if nested_value is not None and nested_value.text is not None:
                            row_data[full_nested_name] = nested_value.text.strip()

                # Добавляем строку в результаты
                rows.append(dict(row_data))

        return rows

    except ET.ParseError as e:
        raise ValueError(f"Ошибка парсинга XML файла: {str(e)}")
    except Exception as e:
        raise RuntimeError(f"Неожиданная ошибка при чтении файла: {str(e)}")


def extract_number(value):
    """Извлекает числовую часть из строки, содержащей буквы и цифры"""
    if pd.isna(value) or value == '':
        return None

    # Ищем все числа в строке (включая числа с плавающей точкой)
    numbers = re.findall(r"[-+]?\d*\.\d+|[-+]?\d+", str(value).replace(',', '.'))

    if numbers:
        # Возвращаем первое найденное число
        try:
            return float(numbers[0])
        except ValueError:
            return None
    return None


def is_analog_value(value):
    """Проверяет, является ли значение аналоговым (содержит буквы)"""
    if pd.isna(value) or value == '':
        return False
    return bool(re.search(r'[a-zA-Z]', str(value)))


def remove_w_prefix(value):
    """Удаляет префикс 'W' из значения, если он присутствует"""
    if pd.isna(value) or value == '':
        return value

    str_value = str(value)
    # Удаляем W в начале строки (без учета регистра)
    if str_value.upper().startswith('W'):
        return str_value[1:]
    return str_value


def _get_filter_column():
    """Выбор столбца для фильтрации с понятными названиями"""
    available_columns = list(CONFIG['COLUMN_DISPLAY_NAMES'].keys())

    print("\nДоступные столбцы для фильтрации:")
    for i, col in enumerate(available_columns, 1):
        print(f"{i}. {CONFIG['COLUMN_DISPLAY_NAMES'][col]}")

    while True:
        try:
            col_choice = input("\nВыберите номер столбца для фильтрации (0 - без фильтра): ")
            col_index = int(col_choice) - 1

            if col_index == -1:
                return None, None
            elif 0 <= col_index < len(available_columns):
                filter_column = available_columns[col_index]
                display_name = CONFIG['COLUMN_DISPLAY_NAMES'][filter_column]
                print(f"\nВыбран столбец для фильтрации: {display_name}")
                return filter_column, display_name
            else:
                print("Неверный выбор. Попробуйте снова.")
        except ValueError:
            print("Пожалуйста, введите число.")


def _get_channel_type():
    """Выбор типа канала для фильтрации по LogicalAddress"""
    print("\nТип канала:")
    print("1. Все каналы")
    print("2. Только дискретные")
    print("3. Только аналоговые")

    while True:
        try:
            type_choice = input("Выберите тип канала (1-3): ")
            if type_choice in ['1', '2', '3']:
                return type_choice
            else:
                print("Неверный выбор. Попробуйте снова.")
        except:
            print("Пожалуйста, введите число от 1 до 3.")


def _get_iotype_filter():
    """Настройка фильтра для столбца IoType"""
    print("\nТипы фильтрации для 'Тип В/В':")
    print("1. Непустые значения")
    print("2. Входа (Input)")
    print("3. Выхода (Output)")

    while True:
        try:
            filter_type = input("Выберите тип фильтрации (1-3): ")
            if filter_type in ['1', '2', '3']:
                break
            else:
                print("Неверный выбор. Попробуйте снова.")
        except:
            print("Пожалуйста, введите число от 1 до 3.")

    if filter_type == '1':
        return "not_empty"
    elif filter_type == '2':
        return "Input"
    elif filter_type == '3':
        return "Output"


def _get_text_filter(filter_column):
    """Настройка фильтра для текстовых столбцов"""
    print("\nТипы фильтрации:")
    print("1. Непустые значения")
    print("2. Конкретное значение")

    while True:
        try:
            filter_type = input("Выберите тип фильтрации (1-2): ")
            if filter_type in ['1', '2']:
                break
            else:
                print("Неверный выбор. Попробуйте снова.")
        except:
            print("Пожалуйста, введите число от 1 до 2.")

    if filter_type == '1':
        return "not_empty"
    elif filter_type == '2':
        return input("Введите значение для фильтрации: ")


def _get_numeric_range_filter():
    """Настройка фильтра по диапазону для номеров каналов"""
    print("\nТипы фильтрации:")
    print("1. Непустые значения")
    print("2. Конкретное значение")
    print("3. Диапазон значений (через тире)")
    print("4. От определенного значения")
    print("5. До определенного значения")

    while True:
        try:
            filter_type = input("Выберите тип фильтрации (1-5): ")
            if filter_type in ['1', '2', '3', '4', '5']:
                break
            else:
                print("Неверный выбор. Попробуйте снова.")
        except:
            print("Пожалуйста, введите число от 1 до 5.")

    if filter_type == '1':
        return "not_empty"
    elif filter_type == '2':
        return input("Введите значение для фильтрации: ")
    elif filter_type in ['3', '4', '5']:
        return _process_range_input(filter_type)


def _process_range_input(filter_type):
    """Обработка ввода диапазона значений"""
    if filter_type == '3':
        prompt = "Введите диапазон значений (например, 10-20): "
    elif filter_type == '4':
        prompt = "Введите начальное значение (например, 10-): "
    elif filter_type == '5':
        prompt = "Введите конечное значение (например, -20): "

    range_input = input(prompt)

    try:
        # Удаляем W из ввода, если присутствует
        range_input = range_input.upper().replace('W', '')

        if filter_type == '3':
            # Полный диапазон
            start, end = map(float, range_input.split('-'))
            return (start, end)
        elif filter_type == '4':
            # От определенного значения
            start = float(range_input.split('-')[0])
            return (start, None)
        elif filter_type == '5':
            # До определенного значения
            end = float(range_input.split('-')[-1])
            return (None, end)
    except ValueError:
        print("Ошибка ввода диапазона. Используется фильтр по непустым значениям.")
        return "not_empty"


def get_filter_settings():
    """Получение настроек фильтра от пользователя с понятными названиями столбцов"""
    # Выбор столбца для фильтрации
    filter_column, display_name = _get_filter_column()
    if filter_column is None:
        return None, None, None

    channel_type = None
    filter_value = None

    # Для столбца "№ канала" добавляем выбор типа канала
    if filter_column == 'LogicalAddress':
        channel_type = _get_channel_type()

    # Выбор типа фильтра в зависимости от столбца
    if filter_column == 'IoType':
        filter_value = _get_iotype_filter()
    elif filter_column in ['InterfaceName', 'Comment']:
        filter_value = _get_text_filter(filter_column)
    elif filter_column == 'LogicalAddress':
        filter_value = _get_numeric_range_filter()

    return filter_column, filter_value, channel_type


def apply_filter(df, filter_column, filter_value, channel_type):
    """Применение фильтра к DataFrame"""
    if filter_column is None:
        return df

    # Проверяем, что столбец для фильтрации существует в DataFrame
    if filter_column not in df.columns:
        print(f"⚠️ Предупреждение: столбец '{filter_column}' не найден в данных. Фильтрация не применена.")
        return df

    # Для столбца "№ канала" применяем предварительную фильтрацию по типу канала
    if filter_column == 'LogicalAddress' and channel_type:
        if channel_type == '2':  # Только дискретные
            df = df[df[filter_column].apply(lambda x: not is_analog_value(x) and extract_number(x) is not None)]
        elif channel_type == '3':  # Только аналоговые
            df = df[df[filter_column].apply(is_analog_value)]

    if filter_value == "not_empty":
        # Фильтр по непустым значениям
        df = df[df[filter_column].notna() & (df[filter_column] != '')]
    elif isinstance(filter_value, tuple) and filter_column == 'LogicalAddress':
        # Для номера канала - диапазон числовых значений
        start, end = filter_value

        # Создаем копию DataFrame перед модификацией
        df = df.copy()

        # Используем .loc для безопасного присвоения значений
        df.loc[:, 'extracted_number'] = df[filter_column].apply(extract_number)

        # Создаем маску для фильтрации в зависимости от типа диапазона
        if start is not None and end is not None:
            # Полный диапазон (от и до)
            numeric_mask = (df['extracted_number'] >= start) & (df['extracted_number'] <= end)
        elif start is not None:
            # Только от определенного значения
            numeric_mask = (df['extracted_number'] >= start)
        elif end is not None:
            # Только до определенного значения
            numeric_mask = (df['extracted_number'] <= end)
        else:
            # Если оба None, оставляем все значения
            numeric_mask = pd.Series([True] * len(df))

        # Применяем фильтр и удаляем временный столбец
        df = df.loc[numeric_mask].drop('extracted_number', axis=1)
    elif isinstance(filter_value, list):
        # Фильтр по списку значений
        df = df[df[filter_column].isin(filter_value)]
    else:
        # Фильтр по конкретному значению
        df = df[df[filter_column] == filter_value]

    return df


def _prepare_dataframe(rows, filter_column, filter_value, channel_type):
    """Подготовка DataFrame с применением фильтров"""
    if not rows:
        return pd.DataFrame()

    # Создаем полный DataFrame из всех строк
    df = pd.DataFrame(rows)

    # Оставляем только нужные колонки (если они существуют)
    source_columns = ['InterfaceName', 'Comment', 'IoType', 'LogicalAddress']
    available_columns = [col for col in source_columns if col in df.columns]

    if not available_columns:
        return pd.DataFrame()

    # Оставляем только нужные колонки
    df = df[available_columns]

    # Применяем фильтр
    df = apply_filter(df, filter_column, filter_value, channel_type)

    if len(df) == 0:
        return df

    # Удаляем префикс 'W' из значений столбца LogicalAddress перед экспортом
    if 'LogicalAddress' in df.columns:
        df['LogicalAddress'] = df['LogicalAddress'].apply(remove_w_prefix)

    # Добавляем сквозную нумерацию
    df['RowNumber'] = range(1, len(df) + 1)

    return df


def _setup_excel_worksheets(wb, sheet_name, total_rows):
    """Настройка листов Excel для записи данных"""
    sheets_needed = math.ceil(total_rows / CONFIG['MAX_ROWS_PER_SHEET'])
    worksheets = []

    for sheet_index in range(sheets_needed):
        current_sheet_name = f"{sheet_name}_{sheet_index + 1}" if sheet_index > 0 else sheet_name

        # Если лист уже существует, используем его
        if current_sheet_name in wb.sheetnames:
            ws = wb[current_sheet_name]
            # Очищаем предыдущие данные на листе (кроме заголовков)
            _clear_worksheet_data(ws)
        else:
            # Создаем копию исходного листа
            if sheet_index == 0:
                if sheet_name not in wb.sheetnames:
                    ws = wb.create_sheet(sheet_name)
                else:
                    ws = wb[sheet_name]
            else:
                source_ws = wb[sheet_name]
                ws = wb.copy_worksheet(source_ws)
                ws.title = current_sheet_name
                # Очищаем данные на скопированном листе
                _clear_worksheet_data(ws)

        worksheets.append((ws, sheet_index))

    return worksheets


def _clear_worksheet_data(worksheet):
    """Очистка данных на листе Excel (кроме заголовков)"""
    for row in range(CONFIG['DATA_START_ROW'], worksheet.max_row + 1):
        for col in ['C'] + list(CONFIG['COLUMN_MAPPING'].values()):
            cell = f"{col}{row}"
            if worksheet[cell].value is not None:
                worksheet[cell].value = None


def _write_data_to_worksheet(worksheet, sheet_data, sheet_index):
    """Запись данных на лист Excel"""
    start_idx = sheet_index * CONFIG['MAX_ROWS_PER_SHEET']

    for local_idx, (_, row) in enumerate(sheet_data.iterrows()):
        excel_row = local_idx + CONFIG['DATA_START_ROW']

        # Добавляем сквозную нумерацию в столбец C
        worksheet[f'C{excel_row}'] = row['RowNumber']

        # Записываем основные данные
        for src_col, excel_col in CONFIG['COLUMN_MAPPING'].items():
            if src_col in sheet_data.columns:
                cell_address = f"{excel_col}{excel_row}"
                worksheet[cell_address] = row[src_col]


def process_and_export_data(rows, template_file, output_file, sheet_name):
    """Обработка данных и экспорт в существующий файл на указанный лист с разбивкой по листам"""
    if not rows:
        print("⚠️ Нет данных для экспорта!")
        return None

    print(f"🔍 Обработка {len(rows)} записей...")

    # Проверяем, есть ли вообще нужные колонки в данных
    first_row = rows[0] if rows else {}
    required_columns = ['InterfaceName', 'Comment', 'IoType', 'LogicalAddress']
    available_columns = [col for col in required_columns if col in first_row]

    if not available_columns:
        print("❌ В данных не найдены необходимые колонки для экспорта")
        return None

    print(f"📊 Найдены колонки: {', '.join(available_columns)}")

    # Получаем настройки фильтра от пользователя
    filter_column, filter_value, channel_type = get_filter_settings()

    # Подготавливаем DataFrame
    df = _prepare_dataframe(rows, filter_column, filter_value, channel_type)

    if len(df) == 0:
        print("ℹ️ Нет данных для экспорта после фильтрации!")
        return None

    # Загружаем шаблонный файл
    if not os.path.exists(template_file):
        print(f"❌ Файл шаблона не найден: {template_file}")
        return None

    # Создаем копию шаблонного файла
    shutil.copy2(template_file, output_file)

    # Загружаем книгу
    wb = load_workbook(output_file)

    # Настраиваем листы для записи
    worksheets = _setup_excel_worksheets(wb, sheet_name, len(df))

    # Записываем данные на каждый лист
    for worksheet, sheet_index in worksheets:
        # Определяем диапазон строк для текущего листа
        start_idx = sheet_index * CONFIG['MAX_ROWS_PER_SHEET']
        end_idx = min((sheet_index + 1) * CONFIG['MAX_ROWS_PER_SHEET'], len(df))
        sheet_data = df.iloc[start_idx:end_idx]

        # Записываем данные на лист
        _write_data_to_worksheet(worksheet, sheet_data, sheet_index)

    # Сохраняем изменения
    wb.save(output_file)

    sheets_needed = math.ceil(len(df) / CONFIG['MAX_ROWS_PER_SHEET'])
    print(f"✅ Успешно экспортировано {len(df)} строк в {sheets_needed} лист(ов)")
    return output_file


if __name__ == "__main__":
    try:
        print("=== AML to Excel Converter v0.83 by @SSergeevitch ===")
        print("=== Tools --- Export CAx data --- Save as .aml file ===")

        # Валидация путей
        aml_file_path = _get_safe_input(
            "Введите путь к AML-файлу (.aml): ",
            lambda x: _validate_file_path(x, "AML-файл")
        )

        output_file = _get_safe_input(
            "Введите имя выходного файла (.xlsx): ",
            lambda x: _validate_file_path(x, "выходной файл", check_exists=False)
        )

        # Использование констант
        template_file = CONFIG['TEMPLATE_FILE']
        sheet_name = CONFIG['DEFAULT_SHEET_NAME']

        print(f"📖 Чтение файла: {aml_file_path}")
        # Парсинг файла
        rows = parse_aml_file(aml_file_path)

        if rows:
            print(f"📊 Найдено {len(rows)} интерфейсов")
            # Обработка и экспорт данных
            result_file = process_and_export_data(rows, template_file, output_file, sheet_name)
            if result_file:
                print(f"🎉 Готово! Файл сохранен: {result_file}")
        else:
            print("ℹ️ Нет интерфейсов для экспорта.")

    except FileNotFoundError as e:
        print(f"❌ Ошибка: {e}")
    except ValueError as e:
        print(f"❌ Ошибка ввода: {e}")
    except Exception as e:
        print(f"❌ Неожиданная ошибка: {e}")